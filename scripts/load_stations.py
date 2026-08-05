"""역 마스터 CSV 적재 (Phase 2 항목 G, D-25).

공공데이터포털의 역 관련 CSV를 `station` 테이블에 넣는다. **정적 참조 데이터를
런타임에 API로 긁지 않는다** — 쿼터·네트워크 의존이 사라지고, 열차 안에서
네트워크가 나빠도 역 목록은 항상 뜬다.

여러 파일을 **순서와 무관하게** 합칠 수 있다. 컬럼이 서로 다른 파일들
(좌표만 있는 것 / 역코드만 있는 것)을 각각 적재하면 역명 기준으로 병합된다
(`storage/stations.upsert`가 COALESCE로 기존 값을 보존한다).

## 쓰는 법

    # 1) CSV를 아무 디렉터리에 모아둔다 (기본값: data/reference/)
    # 2) 먼저 무엇이 적재될지 확인
    uv run python scripts/load_stations.py --dry-run
    # 3) 실제 적재
    uv run python scripts/load_stations.py

    # 파일을 직접 지정할 수도 있다
    uv run python scripts/load_stations.py path/to/역위치정보.csv

## 알려진 소스

- **한국철도공사_역 위치 정보** (data.go.kr 15127532) — 지역본부/역명/위도/경도/출입구개수.
  **역코드 컬럼이 없다** → 좌표 축만 채운다
- 역코드를 담은 CSV (역코드/역명) → 코드 축을 채운다
- 인코딩은 보통 CP949(EUC-KR)다. UTF-8도 자동 판별한다

## 컬럼 매핑

공공데이터 CSV는 같은 뜻의 컬럼명이 파일마다 다르다. 헤더 이름을 **별칭 표**로
맞춘다 — 새 파일에서 못 알아보는 헤더가 나오면 `--dry-run`이 그대로 보여주므로
아래 표에 한 줄 추가하면 된다.
"""

from __future__ import annotations

import argparse
import csv
import io
import re
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.domain.models import KST  # noqa: E402
from app.storage.db import get_conn, init_db  # noqa: E402
from app.storage.stations import (  # noqa: E402
    Station,
    count,
    count_usable,
    count_with_coords,
    fill_coords_only,
    normalize_name,
    upsert,
)

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_DIR = ROOT / "data" / "reference"

USAGE_EPILOG = """\
적재 순서 예 (파일 성격에 따라 플래그가 다르다):

  # 1) 역코드 사전 — 1,255행에 본청·열차소 같은 운영 지점이 섞여 있어 usable=0
  uv run python scripts/load_stations.py data/reference/한국철도공사_철도운영정보_역코드_*.csv

  # 2) 간선 여객역 + 좌표 — 여객역 목록임이 확실하므로 --passenger
  uv run python scripts/load_stations.py --passenger "data/reference/한국철도공사_역 위치 정보_*.csv"

  # 3) 도시철도 좌표 보강 — 1,100행 전국 지하철이라 새 역을 만들면 안 된다
  uv run python scripts/load_stations.py --coords-only data/reference/전체_도시철도역사정보_*.xlsx
"""

# 헤더 별칭 → 논리 필드. 비교는 공백·특수문자를 뺀 소문자로 한다.
ALIASES: dict[str, tuple[str, ...]] = {
    "name": ("역명", "역사명", "역이름", "station", "stationname", "stnnm", "역", "정차역명"),
    # ★ `역번호`는 넣지 않는다. '전국 도시철도역사정보'의 역번호는 `D004` 형식으로
    # 코레일 역코드(`3900023`)와 **다른 체계**다. 여기에 매핑하면 code 컬럼에
    # 두 체계가 섞이고, code_for()가 지하철 코드를 돌려준다.
    "code": ("역코드", "stationcode", "stncd", "stationid"),
    "lat": ("위도", "역위도", "latitude", "lat", "y좌표", "ycoord"),
    "lng": ("경도", "역경도", "longitude", "lng", "lon", "x좌표", "xcoord"),
    # `지역본부`(서울본부 등)는 **노선이 아니다** — 여기에 넣으면 안 된다.
    # upsert의 COALESCE가 먼저 온 값을 지키므로, 지역본부가 line을 차지하면
    # 나중에 시각표의 실제 주운행선명(경부선)이 영영 들어오지 못한다.
    "line": ("주운행선명", "선명", "노선명", "line", "linename", "mrntnm"),
}


def _slug(header: str) -> str:
    return "".join(ch for ch in str(header).lower() if ch.isalnum())


def _candidates(header: str) -> set[str]:
    """헤더 하나에서 비교 후보 여러 개를 만든다.

    공공데이터 CSV는 `역코드(STN_CD)`처럼 **한글명(영문코드)** 형태를 자주 쓴다.
    통째로 슬러그를 만들면(`역코드stncd`) 어느 별칭과도 안 맞으므로,
    괄호 앞과 괄호 안을 따로 떼어 함께 후보로 넣는다.
    """
    text = str(header or "")
    out = {_slug(text)}
    if m := re.match(r"^([^(（]*)[(（]([^)）]*)[)）]\s*$", text.strip()):
        out.add(_slug(m.group(1)))
        out.add(_slug(m.group(2)))
    return {c for c in out if c}


def map_headers(headers: list[str]) -> tuple[dict[str, str], list[str]]:
    """헤더 → 논리 필드 매핑과, 못 알아본 헤더 목록을 돌려준다."""
    mapping: dict[str, str] = {}
    unknown: list[str] = []
    for header in headers:
        cands = _candidates(header)
        for field, names in ALIASES.items():
            if cands & {_slug(n) for n in names}:
                # 같은 필드에 여러 컬럼이 매칭되면 첫 번째만 쓴다
                mapping.setdefault(field, header)
                break
        else:
            unknown.append(header)
    return mapping, unknown


def _read_text(path: Path) -> str:
    """공공데이터 CSV는 대개 CP949다. UTF-8(BOM 포함)도 받아준다."""
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "cp949", "utf-8", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise SystemExit(f"인코딩을 판별할 수 없다: {path}")


def _to_float(value: str | None) -> float | None:
    text = str(value or "").strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _read_records(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    """CSV/XLSX 공통 리더. (헤더, 행 딕셔너리 목록).

    공공데이터는 표준데이터를 XLSX로만 주는 경우가 있다
    (예: 전국_도시철도역사정보). 변환을 사람에게 시키지 않는다.
    """
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            import openpyxl  # noqa: PLC0415
        except ImportError:  # pragma: no cover
            raise SystemExit(
                "XLSX를 읽으려면 openpyxl이 필요하다: uv sync (dev 그룹에 포함)"
            ) from None
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(rows)]
        records = [
            {h: ("" if v is None else str(v)) for h, v in zip(headers, row)} for row in rows
        ]
        wb.close()
        return headers, records

    reader = csv.DictReader(io.StringIO(_read_text(path)))
    return list(reader.fieldnames or []), list(reader)


def parse_rows(path: Path) -> tuple[list[Station], list[str], int]:
    """CSV/XLSX → Station 목록. (역, 못 알아본 헤더, 건너뛴 행 수)."""
    headers, records = _read_records(path)
    mapping, unknown = map_headers(headers)

    if "name" not in mapping:
        raise SystemExit(
            f"{path.name}: 역명 컬럼을 찾을 수 없다.\n"
            f"  헤더: {headers}\n"
            f"  → scripts/load_stations.py 의 ALIASES['name']에 실제 헤더명을 추가하라."
        )

    stations: list[Station] = []
    skipped = 0
    for row in records:
        name = normalize_name(row.get(mapping["name"], ""))
        if not name:
            skipped += 1
            continue
        code = str(row.get(mapping["code"], "") or "").strip() if "code" in mapping else None
        lat = _to_float(row.get(mapping["lat"])) if "lat" in mapping else None
        lng = _to_float(row.get(mapping["lng"])) if "lng" in mapping else None
        line = str(row.get(mapping["line"], "") or "").strip() if "line" in mapping else None
        stations.append(
            Station(name=name, code=code or None, lat=lat, lng=lng, line=line or None)
        )
    return stations, unknown, skipped


def find_code_collisions(stations: list[Station]) -> dict[str, list[str]]:
    """정규화 후 **같은 역명에 서로 다른 역코드**가 붙은 경우를 찾는다.

    역코드 CSV에는 폐역·이설로 구/신 코드가 함께 남아 있다
    (`경주` 3900647/3900895, `동두천` 3900410/3900412 …). 어느 쪽이 현재 쓰이는
    코드인지 CSV만으로는 알 수 없으므로 **임의로 고르지 않고 보고한다.**
    적재는 코드 오름차순의 첫 번째로 **결정적으로** 고정하고, 시각표를 적재할 때
    권위 있는 값으로 덮는다 (`stations.mark_usable(codes=...)`).
    """
    seen: dict[str, set[str]] = {}
    for s in stations:
        if s.code:
            seen.setdefault(s.name, set()).add(s.code)
    return {n: sorted(c) for n, c in seen.items() if len(c) > 1}


def dedupe(stations: list[Station]) -> list[Station]:
    """역명 기준으로 하나만 남긴다. 코드가 여럿이면 **오름차순 첫 번째**.

    '마지막 행이 이긴다'로 두면 파일 정렬이 바뀔 때 저장되는 코드도 바뀐다 —
    재현되지 않는 적재는 디버깅이 불가능하다.
    """
    best: dict[str, Station] = {}
    for s in sorted(stations, key=lambda x: (x.name, x.code or "")):
        prev = best.get(s.name)
        if prev is None:
            best[s.name] = s
            continue
        # 같은 이름의 뒤 행에서는 빈 칸만 메운다 (코드는 첫 번째를 유지)
        best[s.name] = Station(
            name=prev.name,
            code=prev.code or s.code,
            lat=prev.lat if prev.lat is not None else s.lat,
            lng=prev.lng if prev.lng is not None else s.lng,
            line=prev.line or s.line,
            usable=prev.usable or s.usable,
        )
    return [best[k] for k in sorted(best)]


def main() -> None:
    parser = argparse.ArgumentParser(
        description="역 마스터 CSV/XLSX를 station 테이블에 적재한다",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=USAGE_EPILOG,
    )
    parser.add_argument("files", nargs="*", type=Path, help=f"CSV/XLSX (기본: {DEFAULT_DIR}/*)")
    parser.add_argument("--dry-run", action="store_true", help="적재하지 않고 무엇이 들어갈지만 보여준다")
    parser.add_argument(
        "--passenger",
        action="store_true",
        help="이 파일의 역을 여객역으로 확정한다(usable=1). 여객역 목록임이 확실한 파일에만 쓴다",
    )
    parser.add_argument(
        "--coords-only",
        action="store_true",
        help="이미 있는 역의 좌표만 채운다. 새 역을 만들지 않는다 (도시철도 등 넓은 좌표 소스용)",
    )
    args = parser.parse_args()

    if args.passenger and args.coords_only:
        raise SystemExit("--passenger 와 --coords-only 는 함께 쓸 수 없다")

    files = args.files or sorted(
        p for p in DEFAULT_DIR.glob("*") if p.suffix.lower() in (".csv", ".xlsx", ".xlsm")
    )
    if not files:
        raise SystemExit(
            f"CSV가 없다. {DEFAULT_DIR} 에 파일을 넣거나 경로를 인자로 주어라.\n"
            "  예: 한국철도공사_역 위치 정보 (data.go.kr 15127532) 다운로드 후 그 디렉터리에 저장"
        )

    now = datetime.now(KST)
    total_parsed = 0

    with get_conn() as conn:
        init_db()
        before, before_coords = count(conn), count_with_coords(conn)

        for path in files:
            if not path.exists():
                print(f"✗ 없는 파일: {path}")
                continue
            raw_rows, unknown, skipped = parse_rows(path)
            collisions = find_code_collisions(raw_rows)
            stations = dedupe(raw_rows)
            total_parsed += len(raw_rows)
            with_coords = sum(1 for s in stations if s.has_coords)
            with_code = sum(1 for s in stations if s.code)

            print(f"\n── {path.name}")
            print(
                f"   행 {len(raw_rows)}개 (건너뜀 {skipped}) → 역명 {len(stations)}개"
                f" / 좌표 {with_coords} / 역코드 {with_code}"
            )
            if unknown:
                print(f"   ⓘ 못 알아본 헤더 (무시됨): {unknown}")
            for s in stations[:5]:
                print(f"     {s.name:<10} code={s.code or '-':<9} ({s.lat}, {s.lng}) {s.line or ''}")
            if len(stations) > 5:
                print(f"     … 외 {len(stations) - 5}개")

            if collisions:
                print(
                    f"\n   ⚠ 같은 역명에 역코드가 여러 개인 경우 {len(collisions)}건"
                    " (폐역·이설로 구/신 코드가 함께 남은 것)."
                )
                print("     어느 쪽이 현재 코드인지 CSV만으로는 알 수 없어 **오름차순 첫 번째**로")
                print("     결정적으로 고정한다. 시각표를 적재하면 권위 있는 값으로 덮인다.")
                for name, codes in list(collisions.items())[:8]:
                    print(f"       {name:<10} {codes}  → {codes[0]} 채택")
                if len(collisions) > 8:
                    print(f"       … 외 {len(collisions) - 8}건")

            if not args.dry_run:
                if args.coords_only:
                    filled = sum(
                        1 for s in stations if fill_coords_only(conn, s, source=path.name, now=now)
                    )
                    print(f"   → 기존 역의 좌표 {filled}개를 채웠다 (새 역은 만들지 않았다)")
                else:
                    for s in stations:
                        upsert(
                            conn,
                            Station(
                                name=s.name,
                                code=s.code,
                                lat=s.lat,
                                lng=s.lng,
                                line=s.line,
                                usable=args.passenger,
                            ),
                            source=path.name,
                            now=now,
                        )

        if args.dry_run:
            mode = (
                "좌표만 채움(새 역 없음)"
                if args.coords_only
                else ("여객역으로 확정(usable=1)" if args.passenger else "코드 사전(usable=0)")
            )
            print(f"\n[dry-run] 적재하지 않았다. 모드: {mode}. 파싱된 행 {total_parsed}개.")
            print(f"          현재 station 테이블: {before}개 (좌표 {before_coords}개)")
            return

        after, after_coords, usable = count(conn), count_with_coords(conn), count_usable(conn)
        print("\n✅ 적재 완료")
        print(f"   station {before} → {after}개 (좌표 {before_coords} → {after_coords}개)")
        print(f"   여객역 확정(usable) {usable}개 → 드롭다운에 이만큼 노출된다")
        if usable == 0:
            print(
                "\n   ⚠ 아직 여객역이 하나도 확정되지 않았다 — 드롭다운은 Mock 노선으로 폴백한다.\n"
                "     이 파일은 '역코드 사전'이라 여객역과 운영 지점(본청·열차소 등)을\n"
                "     구분할 근거가 없다. 다음 중 하나가 필요하다:\n"
                "       · 한국철도공사_역 위치 정보 (data.go.kr 15127532) 를 함께 적재 → 좌표가 있으면 여객역\n"
                "       · 시각표 정차역 적재 (항목 A 확정 후) → 열차가 서면 여객역"
            )
        elif after - usable:
            print(
                f"   ⓘ 나머지 {after - usable}개는 코드 사전으로만 보관한다"
                " (본청·열차소 등 운영 지점 포함)"
            )


if __name__ == "__main__":
    main()
